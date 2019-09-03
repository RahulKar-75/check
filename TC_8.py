
import os
import sys
import pandas as pd
import time
from openpyxl.chart import (
    LineChart,
    ScatterChart,
    Reference,
    Series
)
import csv
import numpy as np
from openpyxl import load_workbook
from xlsxwriter import Workbook as Workbook
#start=time.time()

def get_avg(list,length):
    sum=0.0
    temp=length*1.0
    for i in range(length):
        sum=sum+list[i]
        
    return sum/temp

def get_max(list,length):
    max=-999.0
    for i in range(length):
        if list[i]>max:
            max=list[i]
            
    return max

def get_min(list,length):
    min=999.0
    for i in range(length):
        if list[i]<min:
            min=list[i]
            
    return min 

def percentile(list,idn):
    a=np.array(list)
    p=np.percentile(a,idn)
    return p

def get_median(list,length):
    
    return np.median(list)


def insert_column(path,sheet_name, column,idx,list,length):
    print(path)
    wb = load_workbook(path)
    ws = wb.get_sheet_by_name(sheet_name)
    
    if idx==0:
        new_column = 11
        
        ws.cell(row=1, column=new_column, value='Iteration')
    elif idx==1:
        new_column = 12
        ws.cell(row=5, column=3, value=float(get_avg(list,length)))
        ws.cell(row=5, column=4, value=float(get_max(list,length)))
        ws.cell(row=5, column=5, value=float(percentile(list,99)))
        ws.cell(row=5, column=6, value=float(get_median(list,length)))
        ws.cell(row=5, column=7, value=float(percentile(list,1)))
        ws.cell(row=5, column=8, value=float(get_min(list,length)))            
        ws.cell(row=1, column=new_column, value='Latency-64GB')
    elif idx==2:
        new_column = 13
        ws.cell(row=8, column=3, value=float(get_avg(list,length)))
        ws.cell(row=8, column=4, value=float(get_max(list,length)))
        ws.cell(row=8, column=5, value=float(percentile(list,99)))
        ws.cell(row=8, column=6, value=float(get_median(list,length)))
        ws.cell(row=8, column=7, value=float(percentile(list,1)))
        ws.cell(row=8, column=8, value=float(get_min(list,length)))  
        ws.cell(row=1, column=new_column, value='Latency-128GB')
    elif idx==3:
        new_column = 14
        ws.cell(row=11, column=3, value=float(get_avg(list,length)))
        ws.cell(row=11, column=4, value=float(get_max(list,length)))
        ws.cell(row=11, column=5, value=float(percentile(list,99)))
        ws.cell(row=11, column=6, value=float(get_median(list,length)))
        ws.cell(row=11, column=7, value=float(percentile(list,1)))
        ws.cell(row=11, column=8, value=float(get_min(list,length)))        
        ws.cell(row=1, column=new_column, value='Latency-256GB')
    for rowy, value in enumerate(column, start=1):
        ws.cell(row=rowy+1, column=new_column, value=value)
        
    wb.save(path)
    #wb.close()

def insert_chart(path,name,idx,length):
    if idx==13:
	print('Creating chart')
        wb = load_workbook(path)
        sheet = wb.get_sheet_by_name(name)      
        chart = ScatterChart(scatterStyle='marker') 
               
        xvalues = Reference(sheet, min_col = 11, 
                            min_row = 2, max_row = int(length-1)) 
        
        yvalues = Reference(sheet, min_col = 12, 
                            min_row = 2, max_row = int(length-1))

        size = Reference(sheet, min_col = 12, 
                     min_row = 2, max_row = int(length-1))
        
        series = Series(values = yvalues, xvalues = xvalues, zvalues=size)   
        chart.series.append(series) 
        chart.title = " 64GB " 
        chart.legend=None
        chart.height = 15
        chart.width = 30
        chart.x_axis.title = " Iteration "
        chart.y_axis.title = " Latency "
        sheet.add_chart(chart, "P2")
        print('Chart created')
               
    elif idx==14:
        wb = load_workbook(path)
        sheet = wb[name]        
        chart = ScatterChart()  
        xvalues = Reference(sheet, min_col = 11, 
                            min_row = 2, max_row = int(length-1)) 
        yvalues = Reference(sheet, min_col = 13, 
                            min_row = 2, max_row = int(length-1))

        size = Reference(sheet, min_col = 13, 
                     min_row = 2, max_row = int(length-1))
        series = Series(values = yvalues, xvalues = xvalues, zvalues=size)
        chart.series.append(series) 
        chart.title = " 128GB "
        chart.legend=None
        chart.height = 15
        chart.width = 30
        chart.x_axis.title = " Iteration "
        chart.y_axis.title = " Latency "
        sheet.add_chart(chart, "P30")
    elif idx==15:
        wb = load_workbook(path)
        sheet = wb[name]        
        chart = ScatterChart()  
        xvalues = Reference(sheet, min_col = 11, 
                            min_row = 2, max_row = int(length-1)) 
        yvalues = Reference(sheet, min_col = 14, 
                            min_row = 2, max_row = int(length-1))
        size = Reference(sheet, min_col = 14, 
                     min_row = 2, max_row = int(length-1))
        series = Series(values = yvalues, xvalues = xvalues, zvalues=size)
        chart.series.append(series) 
        chart.title = " 256GB " 
        chart.legend=None
        chart.height = 15
        chart.width = 30
        chart.x_axis.title = " Iteration "
        chart.y_axis.title = " Latency "
        sheet.add_chart(chart, "P60")    
    
         
    wb.save(path) 
    #wb.close()

def main_insert(iteration,latency,path,capacity):
    path=path.replace('"','')
    res=len(iteration)
    create_flag=False
    #wb = load_workbook(path) 
    
    if int(capacity)==64:
        insert_column(path,'8_FragmentedFileWrite', iteration,0,latency,res)
        insert_column(path,'8_FragmentedFileWrite', latency,1,latency,res)
        insert_chart(path, '8_FragmentedFileWrite', 13,res)  
    elif int(capacity)==128:
        insert_column(path,'8_FragmentedFileWrite', iteration,0,latency,res)
        insert_column(path,'8_FragmentedFileWrite', latency,2,latency,res)
        insert_chart(path, '8_FragmentedFileWrite', 14,res)       
    elif int(capacity)==256:
        insert_column(path,'8_FragmentedFileWrite', iteration,0,latency,res)
        insert_column(path,'8_FragmentedFileWrite', latency,3,latency,res)
        insert_chart(path, '8_FragmentedFileWrite', 15,res)           
    else:
        raise ValueError
    

def ms_to_sec(path,capacity,main_path):
    
    path=path.replace('"','')
    df=pd.read_excel(path)
    y_list=[]
    y_list=(df[' runt'])
    x_list=df[' File_number ']
    for i in range(len(y_list)):
        y_list[i]=str(y_list[i])
        x_list[i]=int(x_list[i])
    
    s_list=[]
    length=min(len(x_list),len(y_list))
    for i in range(length):
        line=y_list[i]
        #params=line.split('m')[0]
        #ms=int(params)
	ms=float(line)
        #s=ms/1000.0
        s=float(ms*1.0)
        s_list.append(s)
    #print 'Hola'
    main_insert(x_list,s_list,main_path,capacity)  
    
        
def convert_to_xlsx(path):
    
    name=path.split('.')
    workbook_path='"'+name[0]+".xlsx"+'"'
    workbook_path=workbook_path.replace('"','')
    workbook = Workbook(workbook_path)   
    path=path.replace('"','')
    csvfile=path
    
    worksheet = workbook.add_worksheet()
    with open(csvfile, 'r') as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader):
            for c, col in enumerate(row):
                worksheet.write(r, c, col)
             
    workbook.close()
    
    return workbook_path


if __name__=="__main__":
    ans=raw_input("Enter file path<space>capacity<main_path>\n")
    ans=ans.split(' ')
    path=ans[0]
    cpty=ans[1]
    main_path=ans[2]
    if '.csv' in path:
        path=convert_to_xlsx(path)
        
    ms_to_sec(path,cpty,main_path)
    #print time.time()-start
